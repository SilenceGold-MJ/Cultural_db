# -*- coding: utf-8 -*-
# !/usr/bin/python3
import openpyxl
import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook

from openpyxl.styles import PatternFill  # 导入填充模块
from framework.logger import Logger
logger = Logger(logger="ExportExcle").getlog()


def ExportExcle(addr,listdata):#新建写入数据
    # 设置文件 mingc
    # 打开文件
    wb = openpyxl.Workbook()  # load_workbook(addr)
    # 创建一张新表
    # ws = wb.create_sheet()
    ws = wb.active
    ws.title = '测试记录'
    # 第一行输入
    sheet_names = wb.sheetnames
    sheet1 = wb['测试记录']  # 打开第一个 sheet 工作表

    csv_head=list(listdata[0].keys())
    del csv_head[0] #删除ID行
    del csv_head[14]#删除色值行
    del csv_head[4]  # 删除时间戳
    column=1
    for i in csv_head:
        sheet1.cell(row=1, column=column, value=i).font = Font(bold=True)  # 实际检索结果
        column=column+1

    #ws.append(csv_head)


    for dic in listdata:
        list_row = list(dic.values())
        color=list_row[15].split('_')
        del list_row[0]  # 删除ID行
        del list_row[14]  # 删除色值行
        del list_row[4]  # 删除时间戳

        for i in range(len(list_row)):
            if 12 == i:#判断结果行#Result
                fille = PatternFill('solid', fgColor=color[0])  # 设置填充颜色为 橙色
                font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=color[1])  # 设置字体样式
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value="").fill = fille  # 序列
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value=list_row[i]).font = font  # 序列
            else:
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value=list_row[i])  # 序列



        #logger.info(list_row)
        #ws.append(list_row)

    wb.save(addr)



def SummaryExcle(addr,listdata):
    title='汇总页'
    # n是判断结果颜色的列
    wb = load_workbook(addr)  # load_workbook(addr)# 打开文件
    # ws = wb.create_sheet()# 创建一张新表
    # ws = wb.active
    # ws.title = '汇总页'
    sheet_names = wb.sheetnames  # 获取所有sheet
    if title in sheet_names:
        sheet1 = wb[title]  # 打开测试记录 sheet 工作表
    else:
        wb.create_sheet(title, 0)  # 插入到最开始的位置
        sheet1 = wb[title]
        Excle_head = list(listdata[0].keys())
        column = 1
        del Excle_head[0]  # 删除ID行
        del Excle_head[12]  # 删除色值行
        del Excle_head[4]  # 删除时间戳
        for i in Excle_head:
            sheet1.cell(row=1, column=column, value=i).font = Font(bold=True)  # 实际检索结果
            column = column + 1



    for dic in listdata:

        list_row = list(dic.values())
        color=list_row[13].split('_')

        del list_row[0]  # 删除ID行
        del list_row[12]  # 删除色值行
        del list_row[4]  # 删除时间戳

        for i in range(len(list_row)):
            if 10 == i:#判断结果行#Result
                fille = PatternFill('solid', fgColor=color[0])  # 设置填充颜色为 橙色
                font = Font(u'宋体', size=11, bold=False, italic=False, strike=False, color=color[1])  # 设置字体样式
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value="").fill = fille  # 序列
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value=list_row[i]).font = font  # 序列
            else:
                sheet1.cell(row=list_row[0] + 1, column=i + 1, value=list_row[i])  # 序列


    wb.save(addr)

def VerticalExcle(addr, dic):  # excle文件路径、输入数据，写入工作表名称，结果颜色列
    title='汇总页'
    n=15
    # n是判断结果颜色的列
    wb = load_workbook(addr)  # load_workbook(addr)# 打开文件
    # ws = wb.create_sheet()# 创建一张新表
    # ws = wb.active
    # ws.title = '汇总页'
    sheet_names = wb.sheetnames  # 获取所有sheet
    sheet1 = wb[title]  # 打开测试记录 sheet 工作表
    list_value = list(dic.values())
    del list_value[0]
    del list_value[3]
    sheet1.cell(1, column=n + 1, value='数据汇总').font = Font(bold=True)
    for i in range(len(list_value)):
        sheet1.cell(row=i + 2, column=n + 1, value=list_value[i])  # 序列
    n=14
    # list_key = list(dic.keys())
    # del list_key[0]
    # del list_key[3]
    list_key = ['测试批次：', '测试版本：', '测试时间：', '文物类型数：','测试次数总数：','通过总数：','失败总数：' ,'准确率达标数：' ,'准确率未达标数：' ,'达标率标准阀值','达标率：' ]
    sheet1.cell(1, column=n + 1, value="名称").font = Font(bold=True)
    for i in range(len(list_key)):
        sheet1.cell(row=i + 2, column=n + 1, value=list_key[i])  # 序列


    # sheet1.cell(len(list_value)+4, column=n + 1, value='于' + datetime.datetime.now().strftime('%Y{y}%m{m}%d{d}').format(y='年', m='月',
    #                                                                                                      d='日')).font = Font(
    #     bold=True)
    wb.save(addr)
